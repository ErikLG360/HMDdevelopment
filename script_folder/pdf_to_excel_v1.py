import pdfplumber
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from rapidfuzz import fuzz
'''
Compiles all the data into one workbook and one sheet but not the correct workbook IE the needed template
'''
pdf_path = input("What is the pdf files name (Include the .pdf) CASE SENSITIVE: > ")
template_excel = input("What is the template excel workbook called (Include the .xlsx) CASE SENSITIVE: >")
new_excel = "item_list_one_sheet.xlsx"
output_template_excel = "filled_template.xlsx"
subcontractor = input("Who is the Subcontractor (CASE SENSITIVE): >")

section_pattern = re.compile(r"^(?:\d+\.\s+)?(?:(?:[A-Z][A-Z\s:/&\-,]+)|(?:[A-Z][a-z]+(?:\s+[A-Z][a-z]+)+))$")
item_pattern = re.compile(
    r'^(?:\d+\s+)?'                                   # (1) optional line number
    r'(.+?)\s+'                                       # (2) description
    r'(\d+(?:[\s,]\d{3})*(?:\.\d+)?)'                 # (3) quantity
    r'(?:\s+([A-Z]{1,10}))?'                          # (4) optional unit
    r'(?:\s*\(?\$?\s*([\d\s,]+(?:\.\d{2})?)\)?\s*)?'  # (5) price 1
    r'(?:\s*\(?\$?\s*([\d\s,]+(?:\.\d{2})?)\)?\s*)?$' # (6) price 2
)

sub_contractor_list = {'D9': ['D','F'], 'I9': ['I', 'K'], 'O9': ['O', 'Q'], 'T9': ['T', 'V'], 'Y9': ['Y', 'AA'], 'AD9': ['AD', 'AF']}


sections = {}

def scan_pdf():
    current_section = None
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            for line in text.splitlines():
                line = line.strip()

                # Detect section headers
                if section_pattern.match(line):
                    current_section = line
                    sections[current_section] = []
                    continue
                # Detects if the line matches the item pattern and saves the matched portions that are available and then saves it to a dictionary with the section as a key and the value as a tuple of the item line
                match = item_pattern.match(line)
                if current_section and match:
                    sections[current_section].append(match.groups())

# Write to an empty Excel workbook
def write_to_empty():
    wb = Workbook()
    ws = wb.active
    ws.append(["Item Description", "Estimated Quantity", "Unit", "Price", "Total Price"])
    #Iterate through the Sections dictionary by using the items() function so it is a string,list of tuple
    for idx, (section, lines) in enumerate(sections.items(), start=1):
        for line in lines:
            ws.append(line)
    wb.save(new_excel)
    print(f"Data saved to {new_excel}")

def extract_numbers(text):
    return re.findall(r'\d+', text)

# Write to Template Excel Sheet
def write_to_temp():
    wb = load_workbook(template_excel)
    ws = wb.active

    contractor_key = ''
    quantity_col = ''
    unit_price_col = ''

    for sub in sub_contractor_list:
        sub_name = ws[sub]
        if subcontractor == sub_name.value:
            contractor_key = sub
            temp_list = sub_contractor_list.get(sub)
            quantity_col = column_index_from_string(temp_list[0])
            unit_price_col = column_index_from_string(temp_list[1])
            break

    if not contractor_key:
        print("ERROR: Subcontractor not found in template.")
        return
    match_count = 0
    # Pre-load Excel descriptions so we don't scan column B thousands of times
    excel_desc_cells = []
    for col in ws.iter_cols(min_col=2, max_col=2, min_row=16, max_row=626):
        for cell in col:
            if cell.value:
                excel_desc_cells.append((cell.row, str(cell.value).strip()))

    # MATCH AGAINST PDF
    for section_name, items in sections.items():
        for item in items:
            desc, qty, unit, price1, price2 = item
            pdf_desc = desc.strip() if desc else ""

            for row, excel_desc in excel_desc_cells:
                #ADDED For Fuzzy matchin

                # Description normalization
                pdf_clean = pdf_desc.lower()
                excel_clean = excel_desc.lower()

                # 1. Block unwanted "contains" matches
                if pdf_clean != excel_clean:
                    if pdf_clean in excel_clean or excel_clean in pdf_clean:
                        continue  # prevents Construction Entrance → Construction Entrance Stone

                # 2. Numeric match protection (prevents 18" → 8")
                pdf_nums = extract_numbers(pdf_desc)
                excel_nums = extract_numbers(excel_desc)
                if pdf_nums and excel_nums and pdf_nums[0] != excel_nums[0]:
                    continue

                score = fuzz.ratio(pdf_desc.lower(), excel_desc.lower())
                if score >= 85:
                #if excel_desc == pdf_desc:
                    ws.cell(row=row, column=quantity_col).value = qty
                    ws.cell(row=row, column=unit_price_col).value = price1
                    match_count += 1
                    print(f"[{match_count}] MATCH: {excel_desc} and {pdf_desc} → Row {row} | Qty={qty} | Price={price1}")
    wb.save(output_template_excel)
    print(f"Data saved to {output_template_excel}")


scan_pdf()
write_to_empty()
write_to_temp()
print("------------DONE------------------")
